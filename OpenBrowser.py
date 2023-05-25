from hashlib import new

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time

wb = load_workbook(filename="C:/Users/8cs/Downloads/inputsk.xlsx")

sheetRange = wb['Sheet1']

usernameStr = 'username'
passwordStr = 'password'
# nipStr = '199507052020121007'
# tmtJabStr = '17-08-2022'
# tglSKStr = '16-08-2022'
# noSKStr = 'Kpts.821/BKPP/01/2022'

browser = webdriver.Chrome('C:/Users/8cs/PycharmProjects/chromedriver.exe')
browser.maximize_window()
browser.get('http://192.168.1.137/plik/dbadmin')

username = browser.find_element(By.NAME, "username")
username.send_keys(usernameStr)
password = browser.find_element(By.NAME, "password")
password.send_keys(passwordStr)
time.sleep(1)
nextButton = browser.find_element(By.XPATH, "/html/body/div/div[2]/form/input[3]")
nextButton.click()

WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, "Jabatan")))
time.sleep(2)
browser.find_element(By.LINK_TEXT, 'Jabatan').click()

# Looping input
# i = 2
# while i < len(sheetRange['A'].value):
for i in range(2, sheetRange.max_row+1):
    # driver.get("https://XXXXXXXXXX")
    # Insert in the form the Name of the person
    cell = "A" + str(i)
    nipStr = sheetRange[cell].value

    # Insert in the form the Age of the person
    cell = "B" + str(i)
    ukerStr = sheetRange[cell].value

    cell = "C" + str(i)
    jabStr = sheetRange[cell].value

    cell = "D" + str(i)
    tmtjabStr = sheetRange[cell].value

    cell = "E" + str(i)
    tglskStr = sheetRange[cell].value

    cell = "F" + str(i)
    noskStr = sheetRange[cell].value

    cell = "G" + str(i)
    ketStr = sheetRange[cell].value

    # driver.find_element_by_xpath("xpath_save_as_draft").click()
    print(nipStr)
# i = 2
#
# while i <= len(sheetRange['A']):
#     nipStr = sheetRange['A'+str(i)].value
#     tmtjabStr = sheetRange['B'+str(i)].value
#     tglskStr = sheetRange['C'+str(i)].value
#     noskStr = sheetRange['D'+str(i)].value


# WebDriverWait(browser,10).until(EC.visibility_of_element_located(By.XPATH, "/html/body/div[2]/div[1]/table[1]/tbody"))

# To click on the WebElement with text as Login you need to induce WebDriverWait for the element_to_be_clickable() and you can use either of the following Locator Strategies:
# WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, "Jabatan")))
# time.sleep(2)
# browser.find_element(By.LINK_TEXT, 'Jabatan').click()

    try:
        nip = browser.find_element(By.NAME, "kata")
        nip.send_keys(nipStr)
        time.sleep(1.5)
        searchButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[1]/div[2]/form/input[3]")
        searchButton.click()

        WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/table/tbody/tr/td[10]/a"))).click()

        wait = WebDriverWait(browser, 10)
        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'kode_unitkerja'))))
        operation_key.select_by_visible_text(ukerStr)

        # browser.find_element(By.NAME, "kode_unitkerja").click()
        # browser.find_element(By.TE, "/html/body/div[2]/div[1]/form/table/tbody/tr[4]/td[2]/select/option[4]").click()

        # browser.find_element(By.NAME, "kd_jabatan5").click()
        # browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/form/table/tbody/tr[8]/td[2]/select/option[84]").click()
        wait2 = WebDriverWait(browser, 10)
        operation_key2 = Select(wait2.until(EC.element_to_be_clickable((By.NAME, 'kd_jabatan5'))))
        operation_key2.select_by_visible_text(jabStr)

        browser.find_element(By.NAME, "tmt_jab").clear()
        browser.find_element(By.NAME, "tmt_jab").send_keys(tmtjabStr)

        browser.find_element(By.NAME, "tgl_sk_jab").clear()
        browser.find_element(By.NAME, "tgl_sk_jab").send_keys(tglskStr)

        browser.find_element(By.NAME, "no_sk_jab").clear()
        browser.find_element(By.NAME, "no_sk_jab").send_keys(noskStr)

        browser.find_element(By.NAME, "keterangan").clear()
        browser.find_element(By.NAME, "keterangan").send_keys(ketStr)

        # browser.implicitly_wait(10)
        time.sleep(1)
        saveButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/form/table/tbody/tr[15]/td/input[1]").click()

        WebDriverWait(browser, 10).until(EC.alert_is_present())
        time.sleep(1.5)
        browser.switch_to.alert.accept()

    except TimeoutException:
        print("ada yang salah boss")
        pass

    time.sleep(1.5)
    # i = i + 1

print("udahan")


# jabButton = browser.findElement(By.XPATH, '//a[@href="logout.php"]')
# jabButton.click()

# jabButton = browser.findElement(By.LINK_TEXT, "Jabatan")
# jabButton.click()

