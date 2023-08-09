from hashlib import new

# AUTOMASI ENTRI DATA ASN

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time

wb = load_workbook(filename="C:/Users/8cs/Downloads/nakespppk.xlsx")

sheetRange = wb['Sheet1']

usernameStr = 'username'
passwordStr = 'password'


browser = webdriver.Chrome('C:/Users/8cs/PycharmProjects/chromedriver4.exe')
browser.maximize_window()
browser.get('http://192.168.40.245/plik/dbadmin')

username = browser.find_element(By.NAME, "username")
username.send_keys(usernameStr)
password = browser.find_element(By.NAME, "password")
password.send_keys(passwordStr)
time.sleep(1)
nextButton = browser.find_element(By.XPATH, "/html/body/div/div[2]/form/input[3]")
nextButton.click()

WebDriverWait(browser, 15).until(EC.element_to_be_clickable((By.LINK_TEXT, "Data Induk ASN")))
time.sleep(2)
browser.find_element(By.LINK_TEXT, 'Data Induk ASN').click()


for i in range(2, sheetRange.max_row+1):

    cell = "A" + str(i)
    unorStr = sheetRange[cell].value

    cell = "B" + str(i)
    jabStr = sheetRange[cell].value

    cell = "C" + str(i)
    tglSkJabStr = sheetRange[cell].value

    cell = "D" + str(i)
    gelarDStr = sheetRange[cell].value

    cell = "E" + str(i)
    namaStr = sheetRange[cell].value

    cell = "F" + str(i)
    gelarBStr = sheetRange[cell].value

    cell = "G" + str(i)
    nipStr = sheetRange[cell].value

    cell = "H" + str(i)
    noSkStr = sheetRange[cell].value

    cell = "I" + str(i)
    noBknStr = sheetRange[cell].value

    cell = "J" + str(i)
    tglBknStr = sheetRange[cell].value

    cell = "K" + str(i)
    tmtStr = sheetRange[cell].value

    cell = "L" + str(i)
    tglSkStr = sheetRange[cell].value

    cell = "M" + str(i)
    pejabatStr = sheetRange[cell].value

    cell = "N" + str(i)
    bupatiStr = sheetRange[cell].value

    cell = "O" + str(i)
    tmptLahirStr = sheetRange[cell].value

    cell = "P" + str(i)
    kabStr = sheetRange[cell].value

    cell = "Q" + str(i)
    provStr = sheetRange[cell].value

    cell = "R" + str(i)
    nikStr = sheetRange[cell].value

    cell = "S" + str(i)
    statusPegStr = sheetRange[cell].value

    # cell = "T" + str(i)
    # bupatiStr = sheetRange[cell].value
    #
    # cell = "U" + str(i)
    # bupatiStr = sheetRange[cell].value

    cell = "V" + str(i)
    tgktPStr = sheetRange[cell].value

    cell = "W" + str(i)
    namaPStr = sheetRange[cell].value

    cell = "X" + str(i)
    tamatStr = sheetRange[cell].value

    cell = "Y" + str(i)
    golStr = sheetRange[cell].value

    cell = "Z" + str(i)
    tmtGolStr = sheetRange[cell].value



    print(nipStr)

    try:

        # KLIK TOMBOL TAMBAH
        tambahButton = browser.find_element(By.XPATH,
                                          "/html/body/div[2]/div[1]/div[1]/div[1]/input").click()
        time.sleep(1.5)
        wait = WebDriverWait(browser, 10)

        # PILIH SELECT UNOR
        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'kode_unitkerja'))))
        operation_key.select_by_visible_text(unorStr)
        wait = WebDriverWait(browser, 10)

        # PILIH SELECT JAFUNG
        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'kd_jabatan5'))))
        operation_key.select_by_visible_text(jabStr)
        wait = WebDriverWait(browser, 10)


        browser.find_element(By.NAME, "tgl_sk_jab").clear()
        browser.find_element(By.NAME, "tgl_sk_jab").send_keys(tmtStr)

        # browser.find_element(By.NAME, "gd").clear()
        # browser.find_element(By.NAME, "gd").send_keys(gelarDStr)

        browser.find_element(By.NAME, "nama").clear()
        browser.find_element(By.NAME, "nama").send_keys(namaStr)

        browser.find_element(By.NAME, "gb").clear()
        browser.find_element(By.NAME, "gb").send_keys(gelarBStr)

        browser.find_element(By.NAME, "nip_baru").clear()
        browser.find_element(By.NAME, "nip_baru").send_keys(nipStr)

        browser.find_element(By.NAME, "nosk_cpns").clear()
        browser.find_element(By.NAME, "nosk_cpns").send_keys(noSkStr)

        browser.find_element(By.NAME, "nobkn_cpns").clear()
        browser.find_element(By.NAME, "nobkn_cpns").send_keys(noBknStr)

        browser.find_element(By.NAME, "tglnobkn_cpns").clear()
        browser.find_element(By.NAME, "tglnobkn_cpns").send_keys(tglBknStr)

        browser.find_element(By.NAME, "tmt_cpns").clear()
        browser.find_element(By.NAME, "tmt_cpns").send_keys(tmtStr)

        browser.find_element(By.NAME, "tglsk_cpns").clear()
        browser.find_element(By.NAME, "tglsk_cpns").send_keys(tglSkStr)

        browser.find_element(By.NAME, "pejpen_cpns").clear()
        browser.find_element(By.NAME, "pejpen_cpns").send_keys(pejabatStr)

        browser.find_element(By.NAME, "jabpen_cpns").clear()
        browser.find_element(By.NAME, "jabpen_cpns").send_keys(bupatiStr)

        browser.find_element(By.NAME, "tempat_lahir").clear()
        browser.find_element(By.NAME, "tempat_lahir").send_keys(tmptLahirStr)

        browser.find_element(By.NAME, "kabupaten").clear()
        browser.find_element(By.NAME, "kabupaten").send_keys(kabStr)

        browser.find_element(By.NAME, "provinsi").clear()
        browser.find_element(By.NAME, "provinsi").send_keys(provStr)

        browser.find_element(By.NAME, "nik").clear()
        browser.find_element(By.NAME, "nik").send_keys(nikStr)

        # PILIH SELECT STATUSPEG
        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'statuspeg'))))
        operation_key.select_by_visible_text(statusPegStr)
        wait = WebDriverWait(browser, 10)

        # PILIH SELECT TINGKAT PENDIDIKAN
        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'kode_pddkn'))))
        operation_key.select_by_visible_text(tgktPStr)
        wait = WebDriverWait(browser, 10)

        browser.find_element(By.NAME, "nama_pddkn").clear()
        browser.find_element(By.NAME, "nama_pddkn").send_keys(namaPStr)

        browser.find_element(By.NAME, "thn_lulus").clear()
        browser.find_element(By.NAME, "thn_lulus").send_keys(tamatStr)

        # PILIH SELECT TINGKAT GOLONGAN
        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'kode_gol'))))
        operation_key.select_by_visible_text(golStr)
        wait = WebDriverWait(browser, 10)

        browser.find_element(By.NAME, "tmt_gol").clear()
        browser.find_element(By.NAME, "tmt_gol").send_keys(tmtGolStr)

        # browser.implicitly_wait(10)
        time.sleep(1)
        saveButton = browser.find_element(By.XPATH,
                                          "/html/body/div[2]/div[1]/form/table/tbody/tr[17]/td/input[1]").click()

        WebDriverWait(browser, 10).until(EC.alert_is_present())
        time.sleep(1.5)
        browser.switch_to.alert.accept()

        # WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/table/tbody/tr/td[12]/a/img"))).click()

        # WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "(//img[@title='Hapus Record Jabatan'])[last()]"))).click()

        time.sleep(1)
        # saveButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/form/table/tbody/tr[15]/td/input[1]").click()
        #
        # WebDriverWait(browser, 10).until(EC.alert_is_present())
        # time.sleep(1.5)
        # browser.switch_to.alert.accept()

    except TimeoutException:
        print("ada yang salah boss")
        pass

    time.sleep(1.5)

print("udahan")


