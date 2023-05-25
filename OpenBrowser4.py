from hashlib import new

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time

wb = load_workbook(filename="D:/BKPP/2023/SKP 2023/4. APRIL/KP APRIL IIID KE BAWAH.xlsx")

sheetRange = wb['Sheet1']

usernameStr = 'username'
passwordStr = 'password'


browser = webdriver.Chrome('C:/Users/8cs/PycharmProjects/chromedriver2.exe')
browser.maximize_window()
browser.get('http://192.168.40.245/plik/dbadmin')

username = browser.find_element(By.NAME, "username")
username.send_keys(usernameStr)
password = browser.find_element(By.NAME, "password")
password.send_keys(passwordStr)
time.sleep(1)
nextButton = browser.find_element(By.XPATH, "/html/body/div/div[2]/form/input[3]")
nextButton.click()

WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, "Golongan")))
time.sleep(2)
browser.find_element(By.LINK_TEXT, 'Golongan').click()


for i in range(2, sheetRange.max_row+1):

    cell = "E" + str(i)
    nipStr = sheetRange[cell].value

    cell = "F" + str(i)
    golStr = sheetRange[cell].value

    cell = "G" + str(i)
    tmtStr = sheetRange[cell].value

    cell = "L" + str(i)
    noskStr = sheetRange[cell].value

    cell = "B" + str(i)
    nobknStr = sheetRange[cell].value

    cell = "C" + str(i)
    tglbknStr = sheetRange[cell].value

    cell = "K" + str(i)
    angkakreditStr = sheetRange[cell].value

    cell = "H" + str(i)
    tahunkerjaStr = sheetRange[cell].value

    cell = "I" + str(i)
    bulankerjaStr = sheetRange[cell].value

    cell = "J" + str(i)
    gajiStr = sheetRange[cell].value

    cell = "M" + str(i)
    tglskStr = sheetRange[cell].value

    cell = "N" + str(i)
    bupatiStr = sheetRange[cell].value


    print(nipStr)

    try:
        nip = browser.find_element(By.NAME, "kata")
        nip.send_keys(nipStr)
        time.sleep(1.5)
        searchButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[1]/div[2]/form/input[3]")
        searchButton.click()

        WebDriverWait(browser, 20).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/table/tbody/tr/td[12]/a/img"))).click()

        wait = WebDriverWait(browser, 10)

        operation_key = Select(wait.until(EC.element_to_be_clickable((By.NAME, 'kode_gol'))))
        operation_key.select_by_visible_text(golStr)

        wait = WebDriverWait(browser, 10)

        # browser.find_element(By.NAME, "kode_unitkerja").click()
        # browser.find_element(By.TE, "/html/body/div[2]/div[1]/form/table/tbody/tr[4]/td[2]/select/option[4]").click()

        # browser.find_element(By.NAME, "kd_jabatan5").click()
        # browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/form/table/tbody/tr[8]/td[2]/select/option[84]").click()
        # wait2 = WebDriverWait(browser, 10)
        # operation_key2 = Select(wait2.until(EC.element_to_be_clickable((By.NAME, 'kd_jabatan5'))))
        # operation_key2.select_by_visible_text(jabStr)

        browser.find_element(By.NAME, "tmt_gol").clear()
        browser.find_element(By.NAME, "tmt_gol").send_keys(tmtStr)

        browser.find_element(By.NAME, "nomor_kp").clear()
        browser.find_element(By.NAME, "nomor_kp").send_keys(noskStr)

        browser.find_element(By.NAME, "nomor_kp").clear()
        browser.find_element(By.NAME, "nomor_kp").send_keys(noskStr)

        browser.find_element(By.NAME, "nomor_bkn").clear()
        browser.find_element(By.NAME, "nomor_bkn").send_keys(nobknStr)

        browser.find_element(By.NAME, "tanggal_bkn").clear()
        browser.find_element(By.NAME, "tanggal_bkn").send_keys(tglbknStr)

        browser.find_element(By.NAME, "angka_kredit").clear()
        browser.find_element(By.NAME, "angka_kredit").send_keys(angkakreditStr)

        browser.find_element(By.NAME, "mk_tahun").clear()
        browser.find_element(By.NAME, "mk_tahun").send_keys(tahunkerjaStr)

        browser.find_element(By.NAME, "mk_bulan").clear()
        browser.find_element(By.NAME, "mk_bulan").send_keys(bulankerjaStr)

        browser.find_element(By.NAME, "gaji_pokok").clear()
        browser.find_element(By.NAME, "gaji_pokok").send_keys(gajiStr)

        browser.find_element(By.NAME, "tanggal_kp").clear()
        browser.find_element(By.NAME, "tanggal_kp").send_keys(tglskStr)

        browser.find_element(By.NAME, "penetapan_kp").clear()
        browser.find_element(By.NAME, "penetapan_kp").send_keys(bupatiStr)

        # browser.implicitly_wait(10)
        time.sleep(1)
        saveButton = browser.find_element(By.XPATH,
                                          "/html/body/div[2]/div[1]/form/table[1]/tbody/tr[12]/td/input[1]").click()

        WebDriverWait(browser, 10).until(EC.alert_is_present())
        time.sleep(1.5)
        browser.switch_to.alert.accept()

        # WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/table/tbody/tr/td[12]/a/img"))).click()

        # WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "(//img[@title='Hapus Record Jabatan'])[last()]"))).click()

        time.sleep(1)
        # saveButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/form/table/tbody/tr[15]/td/input[1]").click()
        #
        # WebDriverWait(browser, 10).until(EC.alert_is_present())
        # time.sleep(1.5)
        # browser.switch_to.alert.accept()

    except TimeoutException:
        print("ada yang salah boss")
        pass

    time.sleep(1.5)

print("udahan")


