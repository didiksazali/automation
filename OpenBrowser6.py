from hashlib import new

# AUTOMASI ENTRI DATA ALAMAT

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time

wb = load_workbook(filename="C:/Users/8cs/Downloads/nakespppk.xlsx")

sheetRange = wb['Sheet1']

usernameStr = 'username'
passwordStr = 'password'


browser = webdriver.Chrome('C:/Users/8cs/PycharmProjects/chromedriver4.exe')
browser.maximize_window()
browser.get('http://192.168.40.245/plik/dbadmin')

username = browser.find_element(By.NAME, "username")
username.send_keys(usernameStr)
password = browser.find_element(By.NAME, "password")
password.send_keys(passwordStr)
time.sleep(1)
nextButton = browser.find_element(By.XPATH, "/html/body/div/div[2]/form/input[3]")
nextButton.click()

WebDriverWait(browser, 15).until(EC.element_to_be_clickable((By.LINK_TEXT, "Update Email-HP-NIK-NPWP-Alamat")))
time.sleep(2)
browser.find_element(By.LINK_TEXT, 'Update Email-HP-NIK-NPWP-Alamat').click()


for i in range(2, sheetRange.max_row+1):

    cell = "A" + str(i)
    nipStr = sheetRange[cell].value

    cell = "B" + str(i)
    alamatStr = sheetRange[cell].value

    print(nipStr)

    try:
        nip = browser.find_element(By.NAME, "kata")
        nip.send_keys(nipStr)
        time.sleep(1.5)
        searchButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[1]/div[2]/form/input[3]")
        searchButton.click()

        time.sleep(1.5)
        wait = WebDriverWait(browser, 10)

        browser.find_element(By.NAME, "alamat_peg").clear()
        browser.find_element(By.NAME, "alamat_peg").send_keys(alamatStr)

        # browser.implicitly_wait(10)
        time.sleep(1)
        saveButton = browser.find_element(By.XPATH,
                                          "/html/body/div[2]/div[1]/table/tbody/tr/td[3]/form/input[11]").click()

        # WebDriverWait(browser, 10).until(EC.alert_is_present())
        # time.sleep(1.5)
        # browser.switch_to.alert.accept()

        time.sleep(5)


    except TimeoutException:
        print("ada yang salah boss")
        pass

    time.sleep(1.5)

print("udahan")


