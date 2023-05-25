from hashlib import new

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time

wb = load_workbook(filename="C:/Users/8cs/Downloads/inputsk.xlsx")

sheetRange = wb['Sheet1']

usernameStr = 'username'
passwordStr = 'password'


browser = webdriver.Chrome('C:/Users/8cs/PycharmProjects/chromedriver.exe')
browser.maximize_window()
browser.get('http://192.168.1.137/plik/dbadmin')

username = browser.find_element(By.NAME, "username")
username.send_keys(usernameStr)
password = browser.find_element(By.NAME, "password")
password.send_keys(passwordStr)
time.sleep(1)
nextButton = browser.find_element(By.XPATH, "/html/body/div/div[2]/form/input[3]")
nextButton.click()

WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.LINK_TEXT, "Jabatan")))
time.sleep(2)
browser.find_element(By.LINK_TEXT, 'Jabatan').click()


for i in range(2, sheetRange.max_row+1):

    cell = "A" + str(i)
    nipStr = sheetRange[cell].value

    cell = "B" + str(i)
    ukerStr = sheetRange[cell].value

    cell = "C" + str(i)
    jabStr = sheetRange[cell].value

    cell = "D" + str(i)
    tmtjabStr = sheetRange[cell].value

    cell = "E" + str(i)
    tglskStr = sheetRange[cell].value

    cell = "F" + str(i)
    noskStr = sheetRange[cell].value

    cell = "G" + str(i)
    ketStr = sheetRange[cell].value

    print(nipStr)

    try:
        nip = browser.find_element(By.NAME, "kata")
        nip.send_keys(nipStr)
        time.sleep(1.5)
        searchButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[1]/div[2]/form/input[3]")
        searchButton.click()

        WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[1]/table/tbody/tr/td[9]/a"))).click()

        WebDriverWait(browser, 20).until(EC.element_to_be_clickable((By.XPATH, "(//img[@title='Hapus Record Jabatan'])[last()]"))).click()

        time.sleep(1)
        # saveButton = browser.find_element(By.XPATH, "/html/body/div[2]/div[1]/form/table/tbody/tr[15]/td/input[1]").click()
        #
        # WebDriverWait(browser, 10).until(EC.alert_is_present())
        # time.sleep(1.5)
        # browser.switch_to.alert.accept()

    except TimeoutException:
        print("ada yang salah boss")
        pass

    time.sleep(1.5)

print("udahan")


